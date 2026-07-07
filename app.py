from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, Response
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_migrate import Migrate, upgrade
from datetime import datetime, date, timedelta, timezone
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
load_dotenv()
import os
import logging
import csv
import io
from decimal import Decimal
from logging.handlers import RotatingFileHandler

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'cambio-esta-key-en-produccion')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV') == 'production'
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
csrf = CSRFProtect(app)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# Logging configurado
if not app.debug:
    if not os.path.exists('logs'):
        os.makedirs('logs')
    file_handler = RotatingFileHandler('logs/peluqueria.log', maxBytes=10240, backupCount=10)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
    ))
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info('Peluquería Ema Lira startup')

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    app.logger.warning(f'404: {request.url}')
    return render_template('errors/404.html', info=INFO), 404

@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f'500: {error}', exc_info=True)
    db.session.rollback()
    try:
        db.create_all()
    except Exception:
        pass
    return render_template('errors/500.html', info=INFO), 500

@app.errorhandler(429)
def ratelimit_error(error):
    app.logger.warning(f'Rate limit exceeded: {request.remote_addr} - {request.url}')
    return render_template('errors/429.html', info=INFO), 429

app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL',
    'sqlite:///peluqueria.db'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
migrate = Migrate(app, db)

_db_ok = False


@app.before_request
def ensure_db():
    global _db_ok
    if not _db_ok:
        try:
            db.create_all()
            _db_ok = True
        except Exception:
            pass


AR_TZ = timezone(timedelta(hours=-3))

@app.context_processor
def inject_now():
    return {
        'now': lambda: datetime.now(AR_TZ),
        'entidades': ENTIDADES_EXPORT,
    }


PROMO = {
    'activo': True,
    'nombre': 'Promo +1',
    'descripcion': 'Si ven\u00eds con uno o m\u00e1s amigos, pagan $10.000 cada uno.',
    'precio': 10000,
    'dias_valido': [1, 2],
}

INFO = {
    'direccion': 'Merced 814',
    'horarios': [
        {'dia': 'Martes', 'horario': '15:00 a 20:00'},
        {'dia': 'Mi\u00e9rcoles a S\u00e1bado', 'horario': '09:00 a 12:00 y 15:00 a 20:00'},
    ],
    'telefono': '5491123456789',
}

ESTADOS_CITA = ['pendiente', 'aceptada', 'completada', 'cancelada']


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)


class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    telefono = db.Column(db.String(20))
    email = db.Column(db.String(100))
    notas = db.Column(db.Text)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    citas = db.relationship('Cita', backref='cliente', lazy=True, cascade='all, delete-orphan')
    pagos = db.relationship('Pago', backref='cliente', lazy=True, cascade='all, delete-orphan')


class Cita(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    fecha = db.Column(db.Date, nullable=False)
    hora = db.Column(db.String(10), nullable=False)
    servicio = db.Column(db.String(100), nullable=False)
    precio = db.Column(db.Numeric(10, 2), default=0)
    estado = db.Column(db.String(20), default='pendiente')
    notas = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Pago(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    concepto = db.Column(db.String(200))
    metodo_pago = db.Column(db.String(50), default='efectivo')
    fecha = db.Column(db.DateTime, default=datetime.now)


class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    precio = db.Column(db.Numeric(10, 2), nullable=False)
    stock = db.Column(db.Integer, default=0)
    categoria = db.Column(db.String(50))


class Gasto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    descripcion = db.Column(db.String(200), nullable=False)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    categoria = db.Column(db.String(50))
    fecha = db.Column(db.DateTime, default=datetime.now)


class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    precio = db.Column(db.Numeric(10, 2), nullable=False, default=0)
    duracion = db.Column(db.Integer, default=30)


def normalizar_telefono(t):
    t = t.replace(' ', '').replace('-', '').replace('+', '').replace('(', '').replace(')', '').replace('.', '')
    if t and not t.startswith('54'):
        t = '54' + t
    return t


def crear_admin_si_no_existe():
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(username='admin')
        db.session.add(admin)
    admin.password_hash = generate_password_hash('admin123')
    db.session.commit()


@app.cli.command('init-db')
def init_db_command():
    """Inicializa la base de datos y crea el admin por defecto."""
    # Ejecutar migraciones pendientes (por si flask db upgrade no corrió en deploy)
    try:
        upgrade()
        print('Migraciones aplicadas.')
    except Exception as e:
        print(f'Error en migraciones (continuando): {e}')
    db.create_all()
    crear_admin_si_no_existe()
    if not Service.query.first():
        db.session.add(Service(nombre='Corte de cabello', precio=12000, duracion=30))
        db.session.commit()
        print('Servicio por defecto creado.')
    print('Base de datos inicializada.')


@app.cli.command('reset-admin')
def reset_admin_command():
    """Fuerza el usuario admin con password admin123 (crea o resetea)."""
    admin = User.query.filter_by(username='admin').first()
    if not admin:
        admin = User(username='admin')
        db.session.add(admin)
    admin.password_hash = generate_password_hash('admin123')
    db.session.commit()
    print('Admin reseteado: admin / admin123')


@app.cli.command('seed-services')
def seed_services_command():
    """Pobla la tabla Service con los 7 servicios base."""
    servicios_base = [
        ('Corte de cabello', 12000, 30),
        ('Corte y barba', 15000, 40),
        ('Barba', 5000, 15),
        ('Tinte', 20000, 60),
        ('Lavado', 4000, 10),
        ('Corte infantil', 8000, 20),
        ('Corte + Barba + Lavado', 22000, 50),
    ]
    for nombre, precio, duracion in servicios_base:
        if not Service.query.filter_by(nombre=nombre).first():
            db.session.add(Service(nombre=nombre, precio=precio, duracion=duracion))
    db.session.commit()
    print('Servicios base creados/verificados.')


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Inici\u00e1 sesi\u00f3n para acceder al panel.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function


def generar_horarios_para_dia(dia_semana):
    if dia_semana in (0, 6):
        return []
    if dia_semana == 1:
        return [f'{h:02d}:{m}' for h in range(15, 20) for m in ['00', '30']]
    slots = []
    for h in range(9, 12):
        for m in ['00', '30']:
            slots.append(f'{h:02d}:{m}')
    for h in range(15, 20):
        for m in ['00', '30']:
            slots.append(f'{h:02d}:{m}')
    return slots


def horarios_disponibles(fecha_str):
    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        return []
    dia_semana = fecha.weekday()
    slots = generar_horarios_para_dia(dia_semana)
    ocupados = [c.hora for c in Cita.query.filter_by(fecha=fecha).filter(
        Cita.estado.in_(['pendiente', 'completada'])).all()]
    disponibles = [h for h in slots if h not in ocupados]
    if fecha == date.today():
        hora_actual = datetime.now().strftime('%H:%M')
        disponibles = [h for h in disponibles if h > hora_actual]
    return disponibles


# --- Autenticaci\u00f3n ---

@app.route('/admin/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session.permanent = True
            session['user_id'] = user.id
            session['username'] = user.username
            flash('Iniciaste sesi\u00f3n correctamente.', 'success')
            return redirect(url_for('admin_dashboard'))
        flash('Usuario o contrase\u00f1a incorrectos.', 'danger')
    return render_template('login.html')


@app.route('/admin/logout')
def logout():
    session.clear()
    flash('Cerraste sesi\u00f3n.', 'info')
    return redirect(url_for('login'))


# --- Landing p\u00fablica ---

@app.route('/health')
def health():
    return jsonify({'status': 'ok'}), 200

@app.route('/')
def landing():
    return render_template('landing.html', promo=PROMO, info=INFO)


# --- Dashboard Admin ---

@app.route('/admin')
@login_required
def admin_dashboard():
    hoy = date.today()
    citas_hoy = Cita.query.filter_by(fecha=hoy).order_by(Cita.hora).all()
    clientes_count = Cliente.query.count()
    productos_count = Producto.query.count()
    ingresos_hoy = db.session.query(db.func.sum(Pago.monto)).filter(
        db.func.date(Pago.fecha) == hoy
    ).scalar() or 0
    gastos_hoy = db.session.query(db.func.sum(Gasto.monto)).filter(
        db.func.date(Gasto.fecha) == hoy
    ).scalar() or 0
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    ingresos_semana = db.session.query(db.func.sum(Pago.monto)).filter(
        db.func.date(Pago.fecha) >= inicio_semana
    ).scalar() or 0
    gastos_semana = db.session.query(db.func.sum(Gasto.monto)).filter(
        db.func.date(Gasto.fecha) >= inicio_semana
    ).scalar() or 0
    return render_template('admin_dashboard.html', citas_hoy=citas_hoy, clientes_count=clientes_count,
                           ingresos_hoy=ingresos_hoy, gastos_hoy=gastos_hoy,
                           ingresos_semana=ingresos_semana, gastos_semana=gastos_semana,
                           productos_count=productos_count)


# --- API para gr\u00e1ficos ---

@app.route('/api/dashboard')
@login_required
def api_dashboard():
    hoy = date.today()
    labels = []
    ingresos_data = []
    gastos_data = []
    for i in range(6, -1, -1):
        d = hoy - timedelta(days=i)
        labels.append(d.strftime('%d/%m'))
        ing = db.session.query(db.func.sum(Pago.monto)).filter(
            db.func.date(Pago.fecha) == d
        ).scalar() or 0
        gas = db.session.query(db.func.sum(Gasto.monto)).filter(
            db.func.date(Gasto.fecha) == d
        ).scalar() or 0
        ingresos_data.append(float(ing))
        gastos_data.append(float(gas))
    servicios = db.session.query(Cita.servicio, db.func.count(Cita.id)).filter(
        Cita.estado != 'cancelada'
    ).group_by(Cita.servicio).all()
    servicios_labels = [s[0] for s in servicios]
    servicios_data = [s[1] for s in servicios]
    return jsonify({
        'labels': labels,
        'ingresos': ingresos_data,
        'gastos': gastos_data,
        'servicios_labels': servicios_labels,
        'servicios_data': servicios_data
    })


@app.route('/api/citas-semana')
@login_required
def api_citas_semana():
    hoy = date.today()
    inicio = hoy - timedelta(days=hoy.weekday())
    fin = inicio + timedelta(days=6)
    citas = Cita.query.filter(Cita.fecha.between(inicio, fin)).filter(
        Cita.estado != 'cancelada').order_by(Cita.fecha, Cita.hora).all()
    dias = ['Lunes', 'Martes', 'Mi\u00e9rcoles', 'Jueves', 'Viernes', 'S\u00e1bado', 'Domingo']
    counts = {i: 0 for i in range(7)}
    for c in citas:
        counts[c.fecha.weekday()] = counts.get(c.fecha.weekday(), 0) + 1
    return jsonify({'labels': dias, 'data': [counts[i] for i in range(7)]})


# --- Zona P\u00fablica ---

@app.route('/agendar')
def agendar():
    try:
        servicios = Service.query.order_by(Service.nombre).all()
    except Exception:
        servicios = []
    return render_template('public/agendar.html', servicios=servicios, promo=PROMO, info=INFO)


@app.route('/api/horarios')
@limiter.limit("30 per minute")
def api_horarios():
    fecha = request.args.get('fecha', '')
    return jsonify({'horarios': horarios_disponibles(fecha)})



@app.route('/agendar/confirmar', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def confirmar_turno():
    if request.method == 'GET':
        return redirect(url_for('agendar'))
    data = request.form
    fecha_str = data.get('fecha')
    hora = data.get('hora')
    servicio_nombre = data.get('servicio')
    nombre = data.get('nombre', '').strip()
    telefono = normalizar_telefono(data.get('telefono', ''))
    email = data.get('email', '').strip()
    try:
        cantidad_personas = int(data.get('cantidad_personas', 1))
    except (ValueError, TypeError):
        flash('Cantidad de personas inválida.', 'danger')
        return redirect(url_for('agendar'))

    if not all([fecha_str, hora, servicio_nombre, nombre]):
        flash('Complet\u00e1 todos los campos obligatorios.', 'danger')
        return redirect(url_for('agendar'))

    if hora not in horarios_disponibles(fecha_str):
        flash('Ese horario ya no est\u00e1 disponible. Eleg\u00ed otro.', 'danger')
        return redirect(url_for('agendar'))

    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Fecha inválida.', 'danger')
        return redirect(url_for('agendar'))

    hoy = date.today()
    if fecha < hoy:
        flash('No podés agendar un turno en el pasado.', 'danger')
        return redirect(url_for('agendar'))
    if fecha == hoy:
        ahora = datetime.now().time()
        try:
            hora_sel = datetime.strptime(hora, '%H:%M').time()
            if hora_sel <= ahora:
                flash('Ese horario ya pasó. Elegí otro.', 'danger')
                return redirect(url_for('agendar'))
        except ValueError:
            pass

    servicio_info = Service.query.filter_by(nombre=servicio_nombre).first()
    dia_semana = fecha.weekday()
    if PROMO['activo'] and dia_semana in PROMO['dias_valido'] and cantidad_personas > 1:
        precio_por_persona = PROMO['precio']
    else:
        precio_por_persona = servicio_info.precio if servicio_info else 0
    precio_total = precio_por_persona * cantidad_personas

    cliente = Cliente.query.filter_by(telefono=telefono).first()
    if not cliente:
        cliente = Cliente(nombre=nombre, telefono=telefono, email=email)
        db.session.add(cliente)
        db.session.flush()
    else:
        if not cliente.nombre and nombre:
            cliente.nombre = nombre

    activos = Cita.query.filter(Cita.cliente_id == cliente.id, Cita.estado.in_(['pendiente', 'aceptada'])).count()
    if activos >= 2:
        flash('Ya ten\u00e9s 2 turnos pendientes. Completalos antes de agendar otro.', 'danger')
        return redirect(url_for('agendar'))

    servicio_texto = servicio_nombre
    if cantidad_personas > 1:
        servicio_texto += f' x{cantidad_personas}'

    cita = Cita(
        cliente_id=cliente.id,
        fecha=fecha,
        hora=hora,
        servicio=servicio_texto,
        precio=precio_total,
        estado='pendiente'
    )
    db.session.add(cita)
    db.session.commit()

    return render_template('public/confirmacion.html', cita=cita, servicio=servicio_info,
                           cantidad=cantidad_personas, precio_persona=precio_por_persona, info=INFO)


# --- Admin: Servicios ---


@app.route('/admin/servicios')
@login_required
def listar_servicios():
    page = request.args.get('page', 1, type=int)
    pagination = Service.query.order_by(Service.nombre).paginate(page=page, per_page=20, error_out=False)
    return render_template('servicios.html', servicios=pagination.items, pagination=pagination)


@app.route('/admin/servicios/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_servicio():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        if not nombre:
            flash('El nombre es obligatorio.', 'danger')
            return render_template('servicio_form.html', servicio=request.form)
        try:
            precio = float(request.form.get('precio', 0))
            duracion = int(request.form.get('duracion', 30))
        except (ValueError, TypeError):
            flash('Precio o duración inválidos.', 'danger')
            return render_template('servicio_form.html', servicio=request.form)
        servicio = Service(nombre=nombre, precio=precio, duracion=duracion)
        db.session.add(servicio)
        db.session.commit()
        flash('Servicio creado correctamente', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_servicios')
        return redirect(next_page)
    return render_template('servicio_form.html', servicio=None)


@app.route('/admin/servicios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_servicio(id):
    servicio = Service.query.get_or_404(id)
    if request.method == 'POST':
        servicio.nombre = request.form.get('nombre', servicio.nombre).strip()
        try:
            servicio.precio = float(request.form.get('precio', 0))
            servicio.duracion = int(request.form.get('duracion', 30))
        except (ValueError, TypeError):
            flash('Precio o duración inválidos.', 'danger')
            return render_template('servicio_form.html', servicio=servicio)
        db.session.commit()
        flash('Servicio actualizado', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_servicios')
        return redirect(next_page)
    return render_template('servicio_form.html', servicio=servicio)


@app.route('/admin/servicios/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_servicio(id):
    servicio = Service.query.get_or_404(id)
    db.session.delete(servicio)
    db.session.commit()
    flash('Servicio eliminado', 'success')
    return redirect(url_for('listar_servicios'))


# --- Admin: Clientes ---

@app.route('/admin/clientes')
@login_required
def listar_clientes():
    page = request.args.get('page', 1, type=int)
    pagination = Cliente.query.order_by(Cliente.nombre).paginate(page=page, per_page=20, error_out=False)
    return render_template('clientes.html', clientes=pagination.items, pagination=pagination)


@app.route('/admin/clientes/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_cliente():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        if not nombre:
            flash('El nombre es obligatorio.', 'danger')
            return render_template('cliente_form.html', cliente=request.form)
        cliente = Cliente(
            nombre=nombre,
            telefono=normalizar_telefono(request.form.get('telefono', '')),
            email=request.form.get('email', ''),
            notas=request.form.get('notas', '')
        )
        db.session.add(cliente)
        db.session.commit()
        flash('Cliente registrado correctamente', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_clientes')
        return redirect(next_page)
    return render_template('cliente_form.html', cliente=None)


@app.route('/admin/clientes/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    if request.method == 'POST':
        cliente.nombre = request.form.get('nombre', cliente.nombre).strip()
        cliente.telefono = normalizar_telefono(request.form.get('telefono', ''))
        cliente.email = request.form.get('email', '')
        cliente.notas = request.form.get('notas', '')
        db.session.commit()
        flash('Cliente actualizado', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_clientes')
        return redirect(next_page)
    return render_template('cliente_form.html', cliente=cliente)


@app.route('/admin/clientes/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    db.session.delete(cliente)
    db.session.commit()
    flash('Cliente eliminado', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_clientes')
    return redirect(next_page)


# --- Admin: Citas ---

@app.route('/admin/citas')
@login_required
def listar_citas():
    page = request.args.get('page', 1, type=int)
    estado_filtro = request.args.get('estado', '')
    q = request.args.get('q', '').strip()

    query = Cita.query

    if estado_filtro in ESTADOS_CITA:
        query = query.filter(Cita.estado == estado_filtro)

    if q:
        query = query.join(Cliente).filter(Cliente.nombre.ilike(f'%{q}%'))

    hoy = date.today()
    if estado_filtro == 'hoy':
        query = query.filter(Cita.fecha == hoy)

    query = query.order_by(Cita.fecha.desc(), Cita.hora)
    pagination = query.paginate(page=page, per_page=20, error_out=False)

    hoy_str = hoy.strftime('%Y-%m-%d')
    counts = {e: Cita.query.filter_by(estado=e).count() for e in ESTADOS_CITA}
    counts['hoy'] = Cita.query.filter(Cita.fecha == hoy).count()
    return render_template('citas.html', citas=pagination.items, pagination=pagination, estado_filtro=estado_filtro, q=q, hoy=hoy_str, counts=counts)


@app.route('/admin/citas/nuevo', methods=['GET', 'POST'])
@login_required
def nueva_cita():
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    if request.method == 'POST':
        cliente_id = request.form.get('cliente_id')
        fecha_str = request.form.get('fecha', '')
        hora = request.form.get('hora', '')
        servicio = request.form.get('servicio', '')
        if not all([cliente_id, fecha_str, hora, servicio]):
            flash('Completá todos los campos obligatorios.', 'danger')
            return redirect(url_for('nueva_cita'))

        if cliente_id == '__new__':
            nuevo_nombre = request.form.get('nuevo_nombre', '').strip()
            nuevo_telefono = request.form.get('nuevo_telefono', '').strip()
            if not nuevo_nombre:
                flash('Ingresá el nombre del nuevo cliente.', 'danger')
                return redirect(url_for('nueva_cita'))
            nuevo_cliente = Cliente(nombre=nuevo_nombre, telefono=nuevo_telefono)
            db.session.add(nuevo_cliente)
            db.session.flush()
            cliente_id = nuevo_cliente.id

        try:
            precio = float(request.form.get('precio', 0) or 0)
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            flash('Fecha o precio inválido.', 'danger')
            return redirect(url_for('nueva_cita'))
        cita = Cita(
            cliente_id=cliente_id,
            fecha=fecha,
            hora=hora,
            servicio=servicio,
            precio=precio,
            estado='aceptada',
            notas=request.form.get('notas', '')
        )
        db.session.add(cita)
        db.session.commit()
        flash('Cita registrada correctamente', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_citas')
        return redirect(next_page)
    servicios = Service.query.order_by(Service.nombre).all()
    return render_template('cita_form.html', cita=None, clientes=clientes, servicios=servicios)


@app.route('/admin/citas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cita(id):
    cita = Cita.query.get_or_404(id)
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    if request.method == 'POST':
        cliente_id = request.form.get('cliente_id', cita.cliente_id)
        if cliente_id == '__new__':
            nuevo_nombre = request.form.get('nuevo_nombre', '').strip()
            nuevo_telefono = request.form.get('nuevo_telefono', '').strip()
            if not nuevo_nombre:
                flash('Ingresá el nombre del nuevo cliente.', 'danger')
                return redirect(url_for('editar_cita', id=id))
            nuevo_cliente = Cliente(nombre=nuevo_nombre, telefono=nuevo_telefono)
            db.session.add(nuevo_cliente)
            db.session.flush()
            cliente_id = nuevo_cliente.id
        cita.cliente_id = cliente_id
        fecha_str = request.form.get('fecha', '')
        cita.hora = request.form.get('hora', cita.hora)
        cita.servicio = request.form.get('servicio', cita.servicio)
        try:
            if fecha_str:
                cita.fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            cita.precio = float(request.form.get('precio', 0) or 0)
        except (ValueError, TypeError):
            flash('Fecha o precio inválido.', 'danger')
            return redirect(url_for('editar_cita', id=id))
        estado = request.form.get('estado', 'pendiente')
        cita.estado = estado if estado in ESTADOS_CITA else 'pendiente'
        cita.notas = request.form.get('notas', '')
        db.session.commit()
        flash('Cita actualizada', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_citas')
        return redirect(next_page)
    servicios = Service.query.order_by(Service.nombre).all()
    return render_template('cita_form.html', cita=cita, clientes=clientes, servicios=servicios)


@app.route('/admin/citas/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_cita(id):
    cita = Cita.query.get_or_404(id)
    db.session.delete(cita)
    db.session.commit()
    flash('Cita eliminada', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_citas')
    return redirect(next_page)


@app.route('/admin/citas/cambiar-estado/<int:id>/<estado>', methods=['POST'])
@login_required
def cambiar_estado_cita(id, estado):
    if estado not in ESTADOS_CITA:
        flash('Estado inv\u00e1lido.', 'danger')
        return redirect(url_for('listar_citas'))
    cita = Cita.query.get_or_404(id)
    cita.estado = estado
    db.session.commit()
    flash(f'Estado cambiado a {estado}', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_citas')
    return redirect(next_page)


from whatsapp_api import enviar_recordatorio


@app.route('/admin/enviar-recordatorio/<int:cita_id>', methods=['POST'])
@login_required
def enviar_recordatorio_cita(cita_id):
    cita = Cita.query.get_or_404(cita_id)
    telefono = cita.cliente.telefono
    if not telefono:
        flash('El cliente no tiene teléfono registrado.', 'warning')
        return redirect(request.referrer or url_for('listar_citas'))
    telefono_limpio = telefono.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
    ok, msg = enviar_recordatorio(
        telefono=telefono_limpio,
        nombre_cliente=cita.cliente.nombre,
        fecha=cita.fecha.strftime('%d/%m'),
        hora=cita.hora,
        servicio=cita.servicio,
    )
    flash(msg, 'success' if ok else 'danger')
    return redirect(request.referrer or url_for('listar_citas'))


@app.route('/api/turnos-pendientes-count')
@login_required
def api_turnos_pendientes_count():
    count = Cita.query.filter_by(estado='pendiente').count()
    return jsonify({'count': count})


@app.route('/admin/turnos-pendientes')
@login_required
def turnos_pendientes():
    citas = Cita.query.filter_by(estado='pendiente').order_by(Cita.fecha, Cita.hora).all()
    return render_template('turnos_pendientes.html', citas=citas)


# --- Admin: Pagos ---

@app.route('/admin/pagos')
@login_required
def listar_pagos():
    page = request.args.get('page', 1, type=int)
    pagination = Pago.query.order_by(Pago.fecha.desc()).paginate(page=page, per_page=20, error_out=False)
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    total = db.session.query(db.func.sum(Pago.monto)).scalar() or 0
    return render_template('pagos.html', pagos=pagination.items, pagination=pagination, clientes=clientes, total=total)


@app.route('/admin/pagos/nuevo', methods=['POST'])
@login_required
def nuevo_pago():
    cliente_id = request.form.get('cliente_id')
    monto = request.form.get('monto', '')
    if not cliente_id or not monto:
        flash('Completá todos los campos obligatorios.', 'danger')
        return redirect(url_for('listar_pagos'))
    try:
        monto = float(monto)
    except (ValueError, TypeError):
        flash('Monto inválido.', 'danger')
        return redirect(url_for('listar_pagos'))
    pago = Pago(
        cliente_id=cliente_id,
        monto=monto,
        concepto=request.form.get('concepto', ''),
        metodo_pago=request.form.get('metodo_pago', 'efectivo')
    )
    db.session.add(pago)
    db.session.commit()
    flash('Pago registrado', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_pagos')
    return redirect(next_page)


@app.route('/admin/pagos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_pago(id):
    pago = Pago.query.get_or_404(id)
    clientes = Cliente.query.order_by(Cliente.nombre).all()
    if request.method == 'POST':
        pago.cliente_id = request.form.get('cliente_id', pago.cliente_id)
        monto = request.form.get('monto', '')
        try:
            pago.monto = float(monto) if monto else pago.monto
        except (ValueError, TypeError):
            flash('Monto inválido.', 'danger')
            return redirect(url_for('editar_pago', id=id))
        pago.concepto = request.form.get('concepto', '')
        pago.metodo_pago = request.form.get('metodo_pago', 'efectivo')
        db.session.commit()
        flash('Pago actualizado', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_pagos')
        return redirect(next_page)
    return render_template('pago_form.html', pago=pago, clientes=clientes)


@app.route('/admin/pagos/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_pago(id):
    pago = Pago.query.get_or_404(id)
    db.session.delete(pago)
    db.session.commit()
    flash('Pago eliminado', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_pagos')
    return redirect(next_page)


# --- Admin: Inventario ---

@app.route('/admin/inventario')
@login_required
def listar_productos():
    page = request.args.get('page', 1, type=int)
    pagination = Producto.query.order_by(Producto.nombre).paginate(page=page, per_page=20, error_out=False)
    return render_template('inventario.html', productos=pagination.items, pagination=pagination)


@app.route('/admin/inventario/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_producto():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        precio = request.form.get('precio', '')
        if not nombre or not precio:
            flash('Nombre y precio son obligatorios.', 'danger')
            return render_template('producto_form.html', producto=request.form)
        try:
            precio = float(precio)
            stock = int(request.form.get('stock', 0) or 0)
        except (ValueError, TypeError):
            flash('Precio o stock inválido.', 'danger')
            return render_template('producto_form.html', producto=request.form)
        producto = Producto(
            nombre=nombre,
            descripcion=request.form.get('descripcion', ''),
            precio=precio,
            stock=stock,
            categoria=request.form.get('categoria', '')
        )
        db.session.add(producto)
        db.session.commit()
        flash('Producto agregado', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_productos')
        return redirect(next_page)
    return render_template('producto_form.html', producto=None)


@app.route('/admin/inventario/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_producto(id):
    producto = Producto.query.get_or_404(id)
    if request.method == 'POST':
        producto.nombre = request.form.get('nombre', producto.nombre).strip()
        monto_precio = request.form.get('precio', '')
        try:
            producto.precio = float(monto_precio) if monto_precio else producto.precio
            producto.stock = int(request.form.get('stock', 0) or 0)
        except (ValueError, TypeError):
            flash('Precio o stock inválido.', 'danger')
            return redirect(url_for('editar_producto', id=id))
        producto.descripcion = request.form.get('descripcion', '')
        producto.categoria = request.form.get('categoria', '')
        db.session.commit()
        flash('Producto actualizado', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_productos')
        return redirect(next_page)
    return render_template('producto_form.html', producto=producto)


@app.route('/admin/inventario/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_producto(id):
    producto = Producto.query.get_or_404(id)
    db.session.delete(producto)
    db.session.commit()
    flash('Producto eliminado', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_productos')
    return redirect(next_page)


# --- Admin: Gastos ---

@app.route('/admin/gastos')
@login_required
def listar_gastos():
    page = request.args.get('page', 1, type=int)
    pagination = Gasto.query.order_by(Gasto.fecha.desc()).paginate(page=page, per_page=20, error_out=False)
    total = db.session.query(db.func.sum(Gasto.monto)).scalar() or 0
    return render_template('gastos.html', gastos=pagination.items, pagination=pagination, total=total)


@app.route('/admin/gastos/nuevo', methods=['POST'])
@login_required
def nuevo_gasto():
    descripcion = request.form.get('descripcion', '').strip()
    monto = request.form.get('monto', '')
    if not descripcion or not monto:
        flash('Descripción y monto son obligatorios.', 'danger')
        return redirect(url_for('listar_gastos'))
    try:
        monto = float(monto)
    except (ValueError, TypeError):
        flash('Monto inválido.', 'danger')
        return redirect(url_for('listar_gastos'))
    gasto = Gasto(
        descripcion=descripcion,
        monto=monto,
        categoria=request.form.get('categoria', '')
    )
    db.session.add(gasto)
    db.session.commit()
    flash('Gasto registrado', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_gastos')
    return redirect(next_page)


@app.route('/admin/gastos/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_gasto(id):
    gasto = Gasto.query.get_or_404(id)
    if request.method == 'POST':
        gasto.descripcion = request.form.get('descripcion', gasto.descripcion).strip()
        monto = request.form.get('monto', '')
        try:
            gasto.monto = float(monto) if monto else gasto.monto
        except (ValueError, TypeError):
            flash('Monto inválido.', 'danger')
            return redirect(url_for('editar_gasto', id=id))
        gasto.categoria = request.form.get('categoria', '')
        db.session.commit()
        flash('Gasto actualizado', 'success')
        next_page = request.form.get('next') or request.referrer or url_for('listar_gastos')
        return redirect(next_page)
    return render_template('gasto_form.html', gasto=gasto)


@app.route('/admin/gastos/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_gasto(id):
    gasto = Gasto.query.get_or_404(id)
    db.session.delete(gasto)
    db.session.commit()
    flash('Gasto eliminado', 'success')
    next_page = request.form.get('next') or request.referrer or url_for('listar_gastos')
    return redirect(next_page)


# --- Exportación de Datos ---

ENTIDADES_EXPORT = {
    'clientes': {
        'nombre': 'Clientes',
        'icono': 'bi-people',
        'columnas': ['ID', 'Nombre', 'Teléfono', 'Email', 'Notas', 'Fecha Registro'],
        'datos': lambda: Cliente.query.order_by(Cliente.nombre).all(),
        'filas': lambda rows: [[c.id, c.nombre, c.telefono or '', c.email or '', c.notas or '',
                                 c.fecha_registro.strftime('%d/%m/%Y %H:%M') if c.fecha_registro else ''] for c in rows],
    },
    'citas': {
        'nombre': 'Citas',
        'icono': 'bi-calendar-check',
        'tiene_fechas': True,
        'columnas': ['ID', 'Cliente', 'Fecha', 'Hora', 'Servicio', 'Precio', 'Estado', 'Notas', 'Creada'],
        'datos': lambda: Cita.query.order_by(Cita.fecha.desc(), Cita.hora).all(),
        'filas': lambda rows: [[c.id, c.cliente.nombre, c.fecha.strftime('%d/%m/%Y'), c.hora,
                                 c.servicio, str(c.precio), c.estado, c.notas or '',
                                 c.created_at.strftime('%d/%m/%Y %H:%M') if c.created_at else ''] for c in rows],
    },
    'pagos': {
        'nombre': 'Pagos',
        'icono': 'bi-cash-coin',
        'tiene_fechas': True,
        'columnas': ['ID', 'Cliente', 'Monto', 'Concepto', 'Método', 'Fecha'],
        'datos': lambda: Pago.query.order_by(Pago.fecha.desc()).all(),
        'filas': lambda rows: [[p.id, p.cliente.nombre, str(p.monto), p.concepto or '', p.metodo_pago or '',
                                 p.fecha.strftime('%d/%m/%Y %H:%M') if p.fecha else ''] for p in rows],
    },
    'gastos': {
        'nombre': 'Gastos',
        'icono': 'bi-cash-stack',
        'tiene_fechas': True,
        'columnas': ['ID', 'Descripción', 'Monto', 'Categoría', 'Fecha'],
        'datos': lambda: Gasto.query.order_by(Gasto.fecha.desc()).all(),
        'filas': lambda rows: [[g.id, g.descripcion, str(g.monto), g.categoria or '',
                                 g.fecha.strftime('%d/%m/%Y %H:%M') if g.fecha else ''] for g in rows],
    },
    'inventario': {
        'nombre': 'Inventario',
        'icono': 'bi-box-seam',
        'columnas': ['ID', 'Nombre', 'Descripción', 'Precio', 'Stock', 'Categoría'],
        'datos': lambda: Producto.query.order_by(Producto.nombre).all(),
        'filas': lambda rows: [[p.id, p.nombre, p.descripcion or '', str(p.precio), p.stock or 0, p.categoria or ''] for p in rows],
    },
    'servicios': {
        'nombre': 'Servicios',
        'icono': 'bi-scissors',
        'columnas': ['ID', 'Nombre', 'Precio', 'Duración (min)'],
        'datos': lambda: Service.query.order_by(Service.nombre).all(),
        'filas': lambda rows: [[s.id, s.nombre, str(s.precio), s.duracion] for s in rows],
    },
}


def exportar_csv(nombre_archivo, columnas, filas):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(columnas)
    for fila in filas:
        writer.writerow(fila)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}'},
    )


def exportar_excel(nombre_archivo, columnas, filas):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Datos'
    ws.append(columnas)
    for fila in filas:
        ws.append(fila)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}'},
    )


def exportar_pdf(nombre_archivo, entidad, columnas, filas):
    from fpdf import FPDF

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, entidad['nombre'], ln=True, align='C')
    pdf.ln(5)

    n = len(columnas)
    usable = 277
    col_w = usable / n if n > 0 else usable

    pdf.set_font('Helvetica', 'B', 9)
    pdf.set_fill_color(44, 62, 80)
    pdf.set_text_color(255, 255, 255)
    for col in columnas:
        pdf.cell(col_w, 8, col, border=1, fill=True, align='C')
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    for i, fila in enumerate(filas):
        if i % 2 == 0:
            pdf.set_fill_color(245, 245, 245)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_font('Helvetica', '', 8)
        for val in fila:
            pdf.cell(col_w, 6, str(val), border=1, fill=True, align='C')
        pdf.ln()

    return Response(
        bytes(pdf.output()),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}'},
    )


@app.route('/admin/exportar')
@login_required
def exportar_hub():
    return render_template('exportar.html', entidades=ENTIDADES_EXPORT)


@app.route('/admin/exportar/<entity>', methods=['POST'])
@login_required
def exportar_entity(entity):
    if entity not in ENTIDADES_EXPORT:
        flash('Entidad no válida.', 'danger')
        return redirect(url_for('exportar_hub'))

    entidad = ENTIDADES_EXPORT[entity]
    formato = request.form.get('formato', 'csv')
    rows = entidad['datos']()
    filas = entidad['filas'](rows)
    columnas = entidad['columnas']
    ar_tz = timezone(timedelta(hours=-3))
    ahora = datetime.now(ar_tz).strftime('%Y-%m-%d')
    nombre_archivo = f'{entity}-{ahora}.{formato}'

    if formato == 'csv':
        return exportar_csv(nombre_archivo, columnas, filas)
    elif formato == 'xlsx':
        return exportar_excel(nombre_archivo, columnas, filas)
    elif formato == 'pdf':
        return exportar_pdf(nombre_archivo, entidad, columnas, filas)
    else:
        flash('Formato no soportado.', 'danger')
        return redirect(url_for('exportar_hub'))


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_ENV') != 'production')
